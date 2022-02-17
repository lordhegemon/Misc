import cbpro
import os
import websocket
import pprint
import ta
import numpy as np
import pandas as pd
import cbpro, time
import openpyxl
import ModuleAgnostic as ma
from datetime import date
public_client = cbpro.PublicClient()

# coinbase setup
# trade_symbol = "ETH-USD"
# traded_quantity = .05
# cbpro_fee = .005
#
# # rsi setup
# rsi_period = 14
# rsi_overbought = 70
# rsi_oversold = 30
#
# # data storage setup
# prices = []
# open_position = False
# total_buys = 0
# total_sells = 0
# order_book = {}
# order_book['rsiBUY'] = []
# order_book['priceBUY'] = []
# order_book['priceSELL'] = []
# order_book['rsiSELL'] = []
# order_book['quantityPriceBUY'] = []
# order_book['quantityPriceSELL'] = []
# order_book['feeBUY'] = []
# order_book['feeSELL'] = []


# inherit from the cbpro websocket client
class myWebsocketClient(cbpro.WebsocketClient):
    def __init__(self, url="wss://ws-feed.pro.coinbase.com", products=None, message_type="subscribe", mongo_collection=None,
                 should_print=True, auth=False, api_key="", api_secret="", api_passphrase="", channels=None):
        super().__init__(url, products, message_type, mongo_collection, should_print, auth, api_key, api_secret, api_passphrase, channels)
        self.message_count = 0
        self.jas = self.loadCsvData()
        self.df, self.data = self.loadCsvData()
        self.data = [i for i in self.data if i[0] != 'VADER']
        print(len(self.data))
        self.coin_tickers = [i[0] + "-USD" for i in self.data] + ['BTC-USD'] + ['ETH-USD']
        print(len(self.coin_tickers))
        self.data_book = []
        self.current_btc = 0
        self.current_ether = 0
        self.used_ticker = []
        self.daily_report = []
        pd.set_option('display.max_columns', None)
        # print(self.df)

    def loadCsvData(self):
        df = pd.read_excel("C:\\Users\\colto\\Google Drive\\crypto\\SatoshiValues.xlsx", dtype='object')
        data = df.to_numpy().tolist()
        return df, data

    def on_open(self):
        self.url = "wss://ws-feed.pro.coinbase.com/"
        self.channels = ["ticker"]
        self.products = self.coin_tickers


    def on_message(self, msg):
        try:
            if msg['product_id'] == 'BTC-USD':
                self.current_btc = float(msg['price'])
            elif msg['product_id'] == 'ETH-USD':
                self.current_ether = float(msg['price'])
            else:
                self.dataProcessor(msg)
        except:
            pass

    def on_close(self):
        print("------------ Results ------------\n")
        self.daily_report = sorted(self.daily_report, key=lambda x: x[1])
        for i in range(len(self.daily_report)):
            print(i, self.daily_report[i])
        self.saveData()


    def dataProcessor(self, msg):
        current_id = msg['product_id'][:-4]
        current_price = float(msg['price'])
        satoshi_value = round((current_price/self.current_btc) * 100000000,4)
        gwei_value = round((current_price/self.current_ether) *  1000000000,4)
        original_sat = self.df[self.df['Crypto'] == current_id].to_numpy().tolist()
        satoshi_ratio = round(satoshi_value/original_sat[-1][1],4)
        gwei_ratio = round(gwei_value/original_sat[-1][2],4)
        if current_id not in self.used_ticker:
            self.used_ticker.append(current_id)
            self.daily_report.append([str(date.today()), current_id, satoshi_value, round(original_sat[-1][-1],4), satoshi_ratio])
            # self.daily_report.append([str(date.today()), current_id, satoshi_value, satoshi_ratio, gwei_value, gwei_ratio])
        # print(len(self.daily_report))
        if len(self.daily_report) == 16:
            wsClient.close()

    def saveData(self):
        wb = openpyxl.load_workbook("C:\\Users\\colto\\Google Drive\\crypto\\SatoshiValues.xlsx", data_only=True)
        for i in self.daily_report:
            print(i)
            ws_main = wb['Original']
            ws_main["F2"] = self.current_btc
            ws_main["G2"] = self.current_ether
            ws = wb[i[1]]
            max_row = ws.max_row+1
            ws["A" + str(max_row)] = i[0]
            ws["B" + str(max_row)] = i[2]
            ws["C" + str(max_row)] = i[-1]
        wb.save("C:\\Users\\colto\\Google Drive\\crypto\\SatoshiValues.xlsx")
wsClient = myWebsocketClient()
wsClient.start()