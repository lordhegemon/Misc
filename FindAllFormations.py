import os
import re
import openpyxl

def main():
    formations = ['1 green river', '11 sego', '3 mahogany oil shale bed', '4 wasatch', 'BMSGW', 'BMSGW ', 'BMSW', 'Bar F (Top Target)', 'Bench', 'Bermude',
                  'Birds Nest', 'Black Shale', 'Black Shale ', 'Blackhawk', 'CR-4', 'CR-6', 'CSLTPK', 'Cane Creek', 'Carb Mkr', 'Castle Peak', 'Castle Peak ', 'Castle peak',
     'Castlegate', 'Chinle', 'Clastic 1', 'Clastic 18', 'Clastic 8', 'Dakota', 'Dakota Sandstone', 'Doug Creek ', 'Douglas Creek', 'Duchesne', 'Duchesne River',
                  'Emery EQ', 'GRDNGLCH', 'GRNRVR', 'GRRV', 'GRTN1', 'Garden Gulch', 'Garden Gulch (TGR3)', 'Garden Gulch TGR3', 'Garden Gulch(TGR3)',
     'Green River', 'Green River (GRRV)', 'Green River (GRTN)', 'Green River (GRTN1)', 'Green River Fouche Mkr', 'Iles', 'L. Green River', 'Lateral TD', 'Long Point',
                  'Long Point Base', 'Long Point Mkr', 'Lower Castle Peak', 'Lower GG', 'Lower Garden Gulch', 'Lower Green River', 'MHGNY',
     'MV Braided', 'Mahogany', 'Mahogany Bench', 'Mahoghany', 'Mahogony Bench', 'Mahoigany', 'Mancos', 'Mancos B', 'Mancos B Target', 'Mesaverde', 'Ostracod',
                  'Salt 12', 'Salt 15', 'Salt 1A', 'Salt 5', 'Sego', 'Surface', 'TD', 'TGR-3', 'TGR3', 'Top Green River', 'Travis Shale', 'Tununk', 'UINTA',
     'UTEBTTE', 'Uinta', 'Uinta ', 'Uintah', 'Upper Blue Gate', 'Upper Blue Gate Mkr', 'Upper Emery', 'Upper GG', 'Upper Garden Gulch', 'Upper Green ?River',
                  'Upper Green River', 'Uteland Butte', 'Uteland Butte ', 'Wasatch', 'Wasatch ', 'White Rim', 'base moderately saline water',
     'base of moderately saline water', 'birds nest', 'black shale', 'bmsgw', 'bmsw', 'castle peak', 'castle peak lp', 'douglas creek', 'douglas creek lp',
                  'duchesne', 'duchesne river', 'garden gulch', 'garden gulch member', 'green river', 'green river (grrv)', 'green river (grtn1)',
     'green river saline', 'intermediate casing set depth', 'l. green river', 'lateral target', 'lateral target td', 'lateral td', 'long point', 'long point mkr',
                  'lower castle peak', 'lower green river', 'lower green river (tgr3)*', 'lower green river lp', 'mahogany', 'mahogany bench',
     'mahogany bench mk', 'surface', 'surface casing set depth', 't.d. (permit)', 'td', 'tgr3', 'ub lp', 'uinta', 'uintah fm', 'upper green river', 'uteland butte',
                  'uteland butte lp', 'uteland\xa0butte', 'wasatch', 'wasatch 4 lp', 'wasatch 5 lp', 'wasatch lp']





    main_path = r"C:\\Work"
    all_data_formations = []
    for root, dirs, files in os.walk(main_path):
        for file in files:
            if file.endswith(".xlsx"):
                parsed_file = os.path.join(root, file)
                try:

                    formation_lst, data_list = openWithXL(parsed_file)
                    all_data_formations = all_data_formations + formation_lst
                except:
                    pass
    all_data_formations = list(set(all_data_formations))
    all_data_formations = sorted(all_data_formations)


def openWithXL(file_path):
    wb = openpyxl.load_workbook(file_path, keep_vba = True, data_only = True)
    ws = wb['Formations']
    formation_lst, depth_lst = [], []
    max_row = ws.max_row
    max_col = ws.max_column
    data_list = []
    formation_col, depth_column, starting_row = findColumnForFormations(max_row, max_col, ws)
    for i in range(starting_row+1, starting_row+15):
        formation, depth = str(ws.cell(row = i, column = formation_col).value), str(ws.cell(row = i, column = depth_column).value)
        if formation != 'None' and depth != 'None':
            formation_lst.append(str(ws.cell(row = i, column = formation_col).value))
            depth_lst.append(str(ws.cell(row = i, column = depth_column).value))
    data_list = list(zip(formation_lst, depth_lst))
    data_list = [list(i) for i in data_list]
    return formation_lst, data_list


def findColumnForFormations(max_row, max_col, ws):
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            cell_val = str(ws.cell(row = i, column = j).value)
            if str(cell_val.strip()) == 'Formation':
                return j, j+1, i

main()

