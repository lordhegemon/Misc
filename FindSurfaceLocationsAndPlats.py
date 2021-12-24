import os
import copy
import LatLon23 as LatLon
import openpyxl
import utm
import xlwings as xw
import ModuleAgnostic
import ModuleAgnostic as ma
import pyodbc
import pandas as pd
import numpy as np
import math
import itertools
import re
import math
import ProcessBHLLocation
import GatherPlatDataSet
from matplotlib import pyplot as plt
import numpy as np


def mainProcess():
    df2 = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length',
                                'Degrees', 'Minutes', 'Seconds', 'Alignment', 'Concatenation', 'North Reference', 'Northing', 'Easting', 'API',
                                'Quadrant', 'FNSL Value', 'FNSL Direction', 'FEWL Value', 'FEWL Direction'])
    df_base = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length',
                                    'Degrees', 'Minutes', 'Seconds', 'Alignment', 'North Reference', 'Concatenation', 'API'])

    pd.set_option('display.max_columns', None)
    pd.options.mode.chained_assignment = None
    conn, cursor = sqlConnect()
    data_lst = parseDatabaseForDataWithSectionsAndSHL(cursor)

    # transformDatabaseQueryToDataframe(data_lst)

    # execute1 = ' select SUBSTRING(API, 0, 11) As API, Wh_sec,[Wh_Twpn],[Wh_Twpd], [Wh_RngN], [Wh_RngD], [Wh_Pm], Wh_X, Wh_Y, [Wh_FtNS], [Wh_Ns], [Wh_FtEW], [Wh_EW]'
    # execute2 = ' from [dbo].[tblAPDLoc]'
    # execute3 = ' where Wh_X IS NOT NULL and Wh_Y IS NOT NULL and Wh_FtNS is not Null and Wh_FtEW is not Null and API is not Null and Wh_Ns is not Null and Wh_EW is not Null and Wh_Sec is not Null and Wh_Twpn is not Null and Wh_Twpd is not Null and Wh_RngD is not Null and Wh_RngN is not Null and Wh_Pm is not Null'
    # execute4 = " "  # " where Zone_Name = 'Surface Location'"
    # data_lst = cursor.execute(execute1 + execute2 + execute3 + execute4)
    # data_lst = fixer(data_lst)
    # data_lst = [list(t) for t in set(tuple(element) for element in data_lst)]

    folder = os.getcwd()
    folder = os.path.join(folder, 'AnchorPoints')
    lst_parsed_df = pd.read_excel(os.path.join(folder, 'parsed_closure.xlsx'), dtype='object')

    lst_parsed_lst = lst_parsed_df.to_numpy().tolist()
    # justConvertSidesToPoints(lst_parsed_lst)
    excel_parsed_for_good_closure_lst, excel_parsed_df = findDataInExistingExcelWithCorrectClosure(folder)
    matcher_lst_1 = matcherDF1(excel_parsed_df, data_lst)
    df_noclosure, df_base_lst = transformDatabaseQueryToDataframe(matcher_lst_1)
    matcher_lst_2 = matcherDF2(lst_parsed_df, data_lst)

    df_parsed_fin, df_parsed_fin_lst = transformDatabaseQueryToDataframe(matcher_lst_2)
    df_merged, df_merged_lst = mergeBothPages(df_base_lst, df_parsed_fin_lst)
    lst = transformData2(df_merged_lst)
    # compareConcLists(data_lst, excel_parsed_df)

    # df_all_lst = df_all.to_numpy().tolist()
    #
    # df_good = pd.read_excel(os.path.join(folder, 'parsed_closure.xlsx'), dtype='object')
    # # df_good = pd.read_csv(os.path.join(folder, 'parsed_closure.csv'), encoding="ISO-8859-1")
    # # df_good = df_good.to_numpy().tolist()
    # df_noclosure = pd.read_csv(os.path.join(folder, 'PlatPointsNoClosure.csv'), encoding="ISO-8859-1")
    # # df_noclosure = df_noclosure.to_numpy().tolist()
    # df_anchor = pd.read_csv(os.path.join(folder, 'AnchorPoints2.csv'), encoding="ISO-8859-1")
    # # df_anchor = df_anchor.to_numpy().tolist()

    # # #
    # # compare_df = compareConcLists(df_good, df_all)
    # # lst = matcherDF(df_all, data_lst)
    # #
    # # for j in lst:
    # #     for i in j:
    # #         for k in range(len(i)):
    #
    #         # new_row = {'Section': i[0],
    #         #            'Township': i[1],
    #         #            'Township Direction': i[2],
    #         #            'Range': i[3],
    #         #            'Range Direction': i[4],
    #         #            'Baseline': i[5],
    #         #            'Side': i[6],
    #         #            'Length': i[7],
    #         #            'Degrees': i[8],
    #         #            'Minutes': i[9],
    #         #            'Seconds': i[10],
    #         #            'Alignment': i[11],
    #         #            'Concatenation': str(i[12]),
    #         #            'North Reference': str(i[13]),
    #         #            'Northing': i[14],
    #         #            'Easting': i[15],
    #         #            'API': str(i[16]),
    #         #            'Quadrant': str(i[17]),
    #         #            'FNSL Value': str(i[18]),
    #         #            'FNSL Direction': str(i[19]),
    #         #            'FEWL Value': str(i[20]),
    #         #            'FEWL Direction': str(i[21])}
    # #         conc = "".join([str(j) for j in i[:7]])
    # #         new_row = {'Section': i[0],
    # #                    'Township': i[1],
    # #                    'Township Direction': i[2],
    # #                    'Range': i[3],
    # #                    'Range Direction': i[4],
    # #                    'Baseline': i[5],
    # #                    'Side': i[6],
    # #                    'Length': i[7],
    # #                    'Degrees': i[8],
    # #                    'Minutes': i[9],
    # #                    'Seconds': i[10],
    # #                    'Alignment': i[11],
    # #                    'Concatenation': i[12],
    # #                    'API': i[13],
    # #                    'Northing': i[14],
    # #                    'Easting': i[15],
    # #                    'FNSL Value': str(i[16]),
    # #                    'FNSL Direction': str(i[17]),
    # #                    'FEWL Value': str(i[18]),
    # #                    'FEWL Direction': str(i[19])}
    #
    # #         df2 = df2.append(new_row, ignore_index=True)
    #
    # # # df2 = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length', 'Degrees', 'Minutes', 'Seconds', 'Alignment', 'North Reference', 'Concatenation', 'API'])
    #
    # df_lst = pd.read_excel(os.path.join(folder, 'AnchorPoints.xlsx'), dtype='object')
    # lst = df_lst.to_numpy().tolist()
    # # df2.to_csv('AnchorPoints2.csv')
    # # transformData(lst)
    # lst = transformData2(lst)

    # for i in lst:
    #     new_row = {'Section': str(i[0]),
    #                'Township': str(i[1]),
    #                'Township Direction': str(i[2]),
    #                'Range': str(i[3]),
    #                'Range Direction': str(i[4]),
    #                'Baseline': str(i[5]),
    #                'Side': str(i[6]),
    #                'Length': str(i[7]),
    #                'Degrees': str(i[8]),
    #                'Minutes': str(i[9]),
    #                'Seconds': str(i[10]),
    #                'Alignment': str(i[11]),
    #                'North Reference': str(i[12]),
    #                'Concatenation': str(i[13]),
    #                'API': str(i[14])}
    #     df_base = df_base.append(new_row, ignore_index=True)
    # df_base.to_csv('PlatPointsNoClosure.csv')


# 162622011

def justConvertSidesToPoints(lst):
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]
    sides_grouped, counter, coord_data_lst, coord_data_lst_grid, shl_coords = [[]], 0, [], [], []
    for i in lst:
        sides_grouped[-1].append(i)
        if counter == 15:
            counter = 0
            sides_grouped.append([])
        else:
            counter += 1
    sides_grouped = [i for i in sides_grouped if i]
    # for i in sides_grouped:
    #     coord_data_lst.append([])
    #     coord_data_lst_grid.append([])
    #     all_data = []
    #     data_converted = sideDataToDecimalAzimuth(dir_lst, i)
    #     dir_lst_flatten = ma.manyToOne(dir_lst)
    #     coordLst = getCoordsLst(data_converted, [0.0, 0.0])
    #     eqLst, coordLst, cornersLst = linesMain(coordLst, dir_lst_flatten)
    #     surfaceLoc = i[0][-4:]
    #     surfaceCoord, xMin, xMax, yMin, yMax = GatherPlatDataSet.getQuad(coordLst, surfaceLoc)
    #     surfaceCoord = [i * 0.3048 for i in surfaceCoord]
    #     shl_coords.append([i[0][16], i[0][17], i[0][13]])
    #     change_x, change_y = abs(i[0][16] - surfaceCoord[0]), abs(i[0][17] - surfaceCoord[1])
    #
    #     for j in coordLst:
    #         for k in j:
    #             coord_data_lst[-1].append([i[0][:6] + [k[0] * 0.3048 + change_x, k[1] * 0.3048 + change_y] + [i[0][13]]])
    #             all_data.append(k)
    #             counter += 1
    #     coord_data_lst[-1] = [k[0] for k in coord_data_lst[-1]]


def parseDatabaseForDataWithSectionsAndSHL(cursor):
    execute1 = ' select Wh_sec,[Wh_Twpn],[Wh_Twpd], [Wh_RngN], [Wh_RngD], [Wh_Pm], NorthReference, iFGridConvergence, X, Y, SUBSTRING(tal.API, 0, 11) As API, [Wh_FtNS], [Wh_Ns], [Wh_FtEW], [Wh_EW]'
    execute2 = ' from [dbo].[DirectionalSurveyHeader] as dsh'
    execute3 = ' join [dbo].[tblAPDLoc] tal on SUBSTRING(tal.API, 0, 11) = dsh.APINumber'
    execute4 = " where Zone_Name = 'Surface Location' and Wh_X IS NOT NULL and Wh_Y IS NOT NULL and Wh_FtNS is not Null and Wh_FtEW is not Null and API is not Null and Wh_Ns is not Null and Wh_EW is not Null and Wh_Sec is not Null and Wh_Twpn is not Null and Wh_Twpd is not Null and Wh_RngD is not Null and Wh_RngN is not Null and Wh_Pm is not Null"

    data_lst = cursor.execute(execute1 + execute2 + execute3 + execute4)
    data_lst = fixer(data_lst)
    data_lst = [list(t) for t in set(tuple(element) for element in data_lst)]
    return data_lst


def findDataInExistingExcelWithCorrectClosure(folder):
    df_all = pd.read_csv(os.path.join(folder, 'PlatGridNumbers.csv'), encoding="ISO-8859-1")
    lst = df_all.to_numpy().tolist()
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]
    good_data = []
    lst = ma.oneToMany(lst, 16)
    for i in lst:

        all_data = []
        data_converted = sideDataToDecimalAzimuth(dir_lst, i)
        dir_lst_flatten = ma.manyToOne(dir_lst)
        coordLst = getCoordsLst(data_converted, [0.0, 0.0])
        eqLst, coordLst, cornersLst = linesMain(coordLst, dir_lst_flatten)
        # surfaceLoc = i[0][-4:]
        # surfaceLoc[0], surfaceLoc[2] = int(float(surfaceLoc[0])), int(float(surfaceLoc[2]))
        for j in coordLst:
            for k in j:
                all_data.append(k)
        start, end = all_data[0], all_data[-1]
        closure_x, closure_y = round(end[0] - start[0], 4), round(end[1] - start[1], 4)
        # if abs(closure_x) > 5 or abs(closure_y) > 5:
        #     counter = 0
        #     for r in range(len(data_converted)):
        #         counter += 1
        # else:
        #     for j in i:
        good_data.append(j)
    return good_data, df_all


def matcherDF1(df, lst):
    found_data = []
    conc_lst = []
    for i in range(len(lst)):
        lst[i][2], lst[i][4], lst[i][5] = translateDirectionToNumber('township', lst[i][2]), translateDirectionToNumber('rng', lst[i][4]), translateDirectionToNumber('baseline', lst[i][5])
        lst[i][6] = lst[i][6][0]
        if lst[i][7] == 'None':
            lst[i][7] = 0
        lesser_conc = str(int(float(lst[i][0]))) + str(int(float(lst[i][1]))) + str(lst[i][2]) + str(int(float(lst[i][3]))) + str(lst[i][4]) + str(lst[i][5])
        re_conc = re.compile(r"^" + re.escape(lesser_conc) + r"\D")
        find = df[(df['Concatenation'].str.contains(re_conc))]

        if len(find) > 0:
            new_line = find.to_numpy().tolist()
            if lesser_conc not in conc_lst:
                for j in range(len(new_line)):
                    new_line[j] = new_line[j][:12] + [new_line[j][-1]]
                    new_line[j] = new_line[j] + [lst[i][10]] + lst[i][6:10] + lst[i][11:] + [lesser_conc]

                    found_data.append(new_line[j])
                    # if lesser_conc == '1632222':
                    #     print(new_line[j])
                conc_lst.append(lesser_conc)

    # ma.printLine(found_data)
    # found_data = ma.oneToMany(found_data, 16)
    d = {}

    for row in found_data:
        # Add name to dict if not exists
        if row[-1] not in d:
            d[row[-1]] = []
        # Add all non-Name attributes as a new list
        d[row[-1]].append(row[:-1])
    for i, k in d.items():
        len(k)
    # ma.printLine(d)

    return found_data


def matcherDF2(df, lst):
    found_data = []
    conc_lst = []
    for i in range(len(lst)):
        lst[i][2], lst[i][4], lst[i][5] = translateDirectionToNumber('township', lst[i][2]), translateDirectionToNumber('rng', lst[i][4]), translateDirectionToNumber('baseline', lst[i][5])
        lst[i][6] = lst[i][6][0]
        if lst[i][5] == 'None':
            lst[i][5] = 0
        lesser_conc = str(int(float(lst[i][0]))) + str(int(float(lst[i][1]))) + lst[i][2] + str(int(float(lst[i][3]))) + lst[i][4] + lst[i][5]
        re_conc = re.compile(r"^" + re.escape(lesser_conc) + r"\D")
        find = df[(df['Concatenation'].str.contains(re_conc))]
        if len(find) > 0:
            new_line = find.to_numpy().tolist()
            if lesser_conc not in conc_lst:
                for j in range(len(new_line)):
                    new_line[j] = new_line[j][:12] + [str(new_line[j][13])] + [str(new_line[j][-1])[:11]]
                    new_line[j] = new_line[j] + lst[i][6:10] + lst[i][11:]
                    found_data.append(new_line[j])
                conc_lst.append(lesser_conc)
    found_data = ma.oneToMany(found_data, 16)
    return found_data


def transformDatabaseQueryToDataframe(lst):
    df_base = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length',
                                    'Degrees', 'Minutes', 'Seconds', 'Alignment', 'Concatenation', 'API', 'North Reference', 'Grid Convergence', 'Northing', 'Easting',
                                    'FNSL Value', 'FNSL Direction', 'FEWL Value', 'FEWL Direction'])
    for j in lst:
        for i in j:
            if i[15] == 'None':
                i[15] = 0
            new_row = {'Section': int(float(i[0])),
                       'Township': int(float(i[1])),
                       'Township Direction': i[2],
                       'Range': int(float(i[3])),
                       'Range Direction': i[4],
                       'Baseline': i[5],
                       'Side': i[6],
                       'Length': i[7],
                       'Degrees': i[8],
                       'Minutes': i[9],
                       'Seconds': i[10],
                       'Alignment': i[11],
                       'Concatenation': i[12],
                       'API': i[13],
                       'North Reference': i[14],
                       'Grid Convergence': float(i[15]),
                       'Northing': float(i[16]),
                       'Easting': float(i[17]),
                       'FNSL Value': float(i[18]),
                       'FNSL Direction': str(i[19]),
                       'FEWL Value': float(i[20]),
                       'FEWL Direction': str(i[21])}

            df_base = df_base.append(new_row, ignore_index=True)
    df_base.to_csv('PlatPointsNoClosure.csv')
    df_base_lst = df_base.to_numpy().tolist()

    return df_base, df_base_lst


def mergeBothPages(lst_parsed, lst_plats):
    df_base = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length',
                                    'Degrees', 'Minutes', 'Seconds', 'Alignment', 'Concatenation', 'API', 'North Reference', 'Grid Convergence', 'Northing', 'Easting',
                                    'FNSL Value', 'FNSL Direction', 'FEWL Value', 'FEWL Direction'])

    conc_lst = []
    merged_lst = []
    lst_parsed = ma.oneToMany(lst_parsed, 16)
    lst_plats = ma.oneToMany(lst_plats, 16)
    lst_all = lst_parsed + lst_plats
    for i in lst_parsed:
        if i[0][12] not in conc_lst:
            conc_lst.append(i[0][12])
            merged_lst.append(i)
    for i in lst_plats:
        if i[0][12] not in conc_lst:
            conc_lst.append(i[0][12])
            merged_lst.append(i)

    for k in lst_all:
        for i in k:
            new_row = {'Section': i[0],
                       'Township': int(float(i[1])),
                       'Township Direction': i[2],
                       'Range': int(float(i[3])),
                       'Range Direction': i[4],
                       'Baseline': i[5],
                       'Side': i[6],
                       'Length': i[7],
                       'Degrees': i[8],
                       'Minutes': i[9],
                       'Seconds': i[10],
                       'Alignment': i[11],
                       'Concatenation': i[12],
                       'API': i[13],
                       'North Reference': i[14],
                       'Grid Convergence': i[15],
                       'Northing': i[16],
                       'Easting': i[17],
                       'FNSL Value': i[18],
                       'FNSL Direction': str(i[19]),
                       'FEWL Value': i[20],
                       'FEWL Direction': str(i[21])}

            df_base = df_base.append(new_row, ignore_index=True)
    # df_base.to_csv('PlatGridNumbers.csv')
    df_base_lst = df_base.to_numpy().tolist()

    return df_base, df_base_lst


def compareConcLists(df_good, df_all):
    dupe_conc_lst = []
    real_all_lst = []
    df_good_conc = df_good['Concatenation'].to_numpy().tolist()
    df_all_conc = df_all['Concatenation'].to_numpy().tolist()
    df_all_lst = df_all.to_numpy().tolist()
    main_list = np.setdiff1d(df_all_conc, df_good_conc)

    for i in range(len(df_all_lst)):
        if df_all_lst[i][-2] not in main_list:
            dupe_conc_lst.append(df_all_lst[i][-2])
            real_all_lst.append(df_all_lst[i])

    find = df_all[(df_all['Concatenation'].isin(main_list))]

    # find = find.to_numpy().tolist()
    # find = ma.oneToMany(find, 16)
    return find


def setDataValues(lst):
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]

    # writeCoordLstSHL = [['West-Up2', 'D25', 'D26', 'D27', 'D28', 'D29'],
    #                     ['West-Up1', 'D33', 'D34', 'D35', 'D36', 'B37'],
    #                     ['West-Down1', 'D41', 'D42', 'D43', 'D44', 'D45'],
    #                     ['West-Down2', 'D49', 'D50', 'D51', 'D52', 'D53'],
    #                     ['East-Up2', 'U25', 'U26', 'U27', 'U28', 'U29'],
    #                     ['East-Up1', 'U33', 'U34', 'U35', 'U36', 'U37'],
    #                     ['East-Down1', 'U41', 'U42', 'U43', 'U44', 'U45'],
    #                     ['East-Down2', 'U49', 'U50', 'U51', 'U52', 'U53'],
    #                     ['North-Left2', 'F18', 'F19', 'F20', 'F21', 'F22'],
    #                     ['North-Left1', 'L18', 'L19', 'L20', 'L21', 'L22'],
    #                     ['North-Right1', 'P18', 'P19', 'P20', 'P21', 'P22'],
    #                     ['North-Right2', 'V18', 'V19', 'V20', 'V21', 'V22'],
    #                     ['South-Left2', 'F57', 'F58', 'F59', 'F60', 'F61'],
    #                     ['South-Left1', 'L57', 'L58', 'L59', 'L60', 'L61'],
    #                     ['South-Right1', 'P57', 'P58', 'P59', 'P60', 'P61'],
    #                     ['South-Right2', 'V57', 'V58', 'V59', 'V60', 'V61']]
    # wb = xw.Book("C:\\Work\\Test scripts\\AnchorPoints\\Casing_Review.xlsx")
    # ws = wb.sheets['SHL Section']
    # # wb = openpyxl.load_workbook("C:\\Work\\Test scripts\\AnchorPoints\\Casing_Review.xlsx", keep_vba=False, data_only=True)
    # # ws = wb['SHL Section']
    # ws.range('N7').value, ws.range('O7').value, ws.range('P7').value, ws.range('Q7').value, ws.range('R7').value, ws.range('S7').value = lst[0][0], lst[0][1], lst[0][2], lst[0][3], lst[0][4], lst[0][5]
    # # ws['N10'], ws['O10'], ws['P10'], ws['Q10'], ws['R10'], ws['S10'] = lst[0][0], lst[0][1], lst[0][2], lst[0][3], lst[0][4], lst[0][5]
    # for i in lst:
    #     val = i[6]
    #     for j in writeCoordLstSHL:
    #         if val == j[0]:
    #             counter = 0
    #             for k in j[1:]:
    #                 ws.range(k).value = float(i[7+counter])
    #
    #                 counter += 1


def transformData(lst):
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]
    good_data = []
    for i in lst:
        all_data = []
        data_converted = sideDataToDecimalAzimuth(dir_lst, i)
        dir_lst_flatten = ma.manyToOne(dir_lst)
        coordLst = getCoordsLst(data_converted, [0.0, 0.0])
        eqLst, coordLst, cornersLst = linesMain(coordLst, dir_lst_flatten)
        surfaceLoc = i[0][-4:]
        surfaceLoc[0], surfaceLoc[2] = int(float(surfaceLoc[0])), int(float(surfaceLoc[2]))
        surfaceCoord, xMin, xMax, yMin, yMax = GatherPlatDataSet.getQuad(coordLst, surfaceLoc)

        for j in coordLst:
            for k in j:
                all_data.append(k)
        start, end = all_data[0], all_data[-1]
        closure_x, closure_y = round(end[0] - start[0], 4), round(end[1] - start[1], 4)
        if abs(closure_x) > 5 or abs(closure_y) > 5:
            counter = 0
            for r in range(len(data_converted)):
                if not checkProximalValues(data_converted[r][1]):
                    if data_converted[r][0] > 0:
                        print('wrong', data_converted[r])

                if not checkProximalValues(data_converted[r][1]):
                    if data_converted[r][0] > 0:
                        print('wrong', data_converted[r])
                if data_converted[r][0] % 1320 == 0 and data_converted[r][0] > 0:
                    print('GLO value', data_converted[r])
                counter += 1
        else:
            for j in i:
                good_data.append(j)


def transformData2(lst):
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]
    good_data = []

    lst = ma.oneToMany(lst, 16)

    coord_data_lst = []
    coord_data_lst_grid = []
    shl_coords = []
    for i in lst:
        coord_data_lst.append([])
        coord_data_lst_grid.append([])
        all_data = []
        data_converted = sideDataToDecimalAzimuth(dir_lst, i)
        dir_lst_flatten = ma.manyToOne(dir_lst)
        try:
            coordLst = getCoordsLst(data_converted, [0.0, 0.0])
            eqLst, coordLst, cornersLst = linesMain(coordLst, dir_lst_flatten)
            surfaceLoc = i[0][-4:]
            surfaceCoord, xMin, xMax, yMin, yMax = GatherPlatDataSet.getQuad(coordLst, surfaceLoc)
            surfaceCoord = [i * 0.3048 for i in surfaceCoord]
            shl_coords.append([i[0][16], i[0][17], i[0][13]])
            change_x, change_y = abs(i[0][16] - surfaceCoord[0]), abs(i[0][17] - surfaceCoord[1])

            counter = 0
            if i[0][14].lower() == 't':
                for j in coordLst:
                    for k in j:
                        coord_data_lst[-1].append([i[0][:6] + [k[0] * 0.3048 + change_x, k[1] * 0.3048 + change_y] + [i[0][13]]])
                        all_data.append(k)
                        counter += 1
                coord_data_lst[-1] = [k[0] for k in coord_data_lst[-1]]
            else:
                for j in coordLst:
                    for k in j:
                        coord_data_lst_grid[-1].append([i[0][:6] + [k[0] * 0.3048 + change_x, k[1] * 0.3048 + change_y] + [i[0][13]]])
                        all_data.append(k)
                        counter += 1
                coord_data_lst_grid[-1] = [k[0] for k in coord_data_lst_grid[-1]]
        except:
            pass

    coord_data_lst = [i for i in coord_data_lst if i]
    coord_data_lst_grid = [i for i in coord_data_lst_grid if i]
    all_data = coord_data_lst + coord_data_lst_grid
    # for i in range(len(coord_data_lst)):
    #     start, end = coord_data_lst[i][0][6:8], coord_data_lst[i][-1][6:8]
    #     closure_x, closure_y = round(end[0] - start[0], 4), round(end[1] - start[1], 4)
    #     if abs(closure_x) > 1 or abs(closure_y) > 1:
    #         coord_data_lst[i][-1] = coord_data_lst[i][0]
    # for i in range(len(coord_data_lst_grid)):
    #     start, end = coord_data_lst_grid[i][0][6:8], coord_data_lst_grid[i][-1][6:8]
    #     closure_x, closure_y = round(end[0] - start[0], 4), round(end[1] - start[1], 4)
    #     if abs(closure_x) > 1 or abs(closure_y) > 1:
    #         coord_data_lst_grid[i][-1] = coord_data_lst_grid[i][0]

    zp = ZoomPan()
    # fig, ax = plt.subplots()
    # figZoom, figPan = zp.zoom_factory(ax, base_scale=1.1), zp.pan_factory(ax)

    shl_coords = [list(t) for t in set(tuple(element) for element in shl_coords)]
    all_x = [j[0] for j in shl_coords]
    all_y = [j[1] for j in shl_coords]
    all_text = [j[2] for j in shl_coords]

    # ax.scatter([surfaceCoord[0]], [surfaceCoord[1]], c='black')
    coordLst_unmerged = list(itertools.chain.from_iterable(coord_data_lst))
    coord_data_lst_grid_unmerged = list(itertools.chain.from_iterable(coord_data_lst_grid))
    all_unmerged = list(itertools.chain.from_iterable(all_data))
    # saveCoordData(all_unmerged)
    return all_data


def saveCoordData(lst):
    df = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Easting', 'Northing', 'API'])
    for i in lst:
        new_row = {'Section': i[0],
                   'Township': int(float(i[1])),
                   'Township Direction': i[2],
                   'Range': int(float(i[3])),
                   'Range Direction': i[4],
                   'Baseline': i[5],
                   'Easting': i[6],
                   'Northing': i[7],
                   'API': i[8]}
        df = df.append(new_row, ignore_index=True)
    df.to_csv('AllGrids.csv')


def checkProximalValues(data):
    check_true = [False, False, False, False]
    if 275 > data > 265:
        check_true[0] = True
    if 185 > data > 175:
        check_true[1] = True
    if 95 > data > 85:
        check_true[2] = True
    if data > 360 or data < 5:
        check_true[3] = True
    if True not in check_true:
        return False
    return True


def sideDataToDecimalAzimuth(dir_lst, data):
    dir_lst_flatten = ma.manyToOne(dir_lst)
    if len(data[0]) == 14:
        new_data = [i[7:12] for i in data]
        new_data = [[float(j) for j in i] for i in new_data]
        new_data = [[data[i][6]] + new_data[i] for i in range(len(new_data))]
        data_converted = convertDirections(new_data, dir_lst_flatten)
    elif len(data[0]) == 22 and data[0][14] == 'T':
        new_data = [i[7:12] for i in data]
        new_data = [[float(j) for j in i] for i in new_data]
        new_data = [[data[i][6]] + new_data[i] for i in range(len(new_data))]
        data_converted = convertDirections(new_data, dir_lst_flatten)

    else:
        new_data = [i[7:12] for i in data]
        new_data = [[float(j) for j in i] for i in new_data]
        for r in range(len(new_data)):
            new_data[r] = new_data[r] + data[r][14:16]
        new_data = [[data[i][6]] + new_data[i] for i in range(len(new_data))]

        data_converted = convertDirections2(new_data, dir_lst_flatten)

    return data_converted


def sideDataToDecimalAzimuth2(dir_lst, data):
    dir_lst_flatten = ma.manyToOne(dir_lst)

    new_data = [i[7:12] for i in data]
    new_data = [[float(j) for j in i] for i in new_data]
    new_data = [[data[i][6]] + new_data[i] for i in range(len(new_data))]
    data_converted = convertDirections(new_data, dir_lst_flatten)

    return data_converted


def gatherValData(data, dir_lst):
    dir_lst_flatten = ma.manyToOne(dir_lst)
    new_data = [i[7:12] for i in data]
    new_data = [[float(j) for j in i] for i in new_data]
    new_data = [[data[i][6]] + new_data[i] for i in range(len(new_data))]
    data_converted = convertDirections2(new_data, dir_lst_flatten)
    return dir_lst_flatten, new_data


def convertDirections(new_data, dir_lst):
    data_converted = []
    for i in range(len(dir_lst)):
        side, deg, min, sec, dir_val = new_data[i][1], new_data[i][2], new_data[i][3], new_data[i][4], new_data[i][5]
        if i > 8:
            if dir_val == 4 or dir_val == 1:
                decVal = 180 - (deg + min / 60 + sec / 3600)
            else:
                decVal = (deg + min / 60 + sec / 3600)
        else:
            if dir_val == 4 or dir_val == 1:
                decVal = 180 - (deg + min / 60 + sec / 3600)
            else:
                decVal = 180 + (deg + min / 60 + sec / 3600)
        data_converted.append([side, decVal])

    data_converted = [data_converted[12], data_converted[13], data_converted[14], data_converted[15],
                      data_converted[7], data_converted[6], data_converted[5], data_converted[4],
                      data_converted[11], data_converted[10], data_converted[9], data_converted[8],
                      data_converted[0], data_converted[1], data_converted[2], data_converted[3]]
    return data_converted


def convertDirections2(new_data, dir_lst):
    data_converted = []
    for i in range(len(dir_lst)):
        side, deg, min, sec, dir_val = new_data[i][1], new_data[i][2], new_data[i][3], new_data[i][4], new_data[i][5]
        if i > 8:
            if dir_val == 4 or dir_val == 1:
                decVal = 180 - (deg + min / 60 + sec / 3600)
            else:
                decVal = (deg + min / 60 + sec / 3600)
        else:
            if dir_val == 4 or dir_val == 1:
                decVal = 180 - (deg + min / 60 + sec / 3600)
            else:
                decVal = 180 + (deg + min / 60 + sec / 3600)
        decVal = decVal + new_data[i][-1]
        data_converted.append([side, decVal])

    data_converted = [data_converted[12], data_converted[13], data_converted[14], data_converted[15],
                      data_converted[7], data_converted[6], data_converted[5], data_converted[4],
                      data_converted[11], data_converted[10], data_converted[9], data_converted[8],
                      data_converted[0], data_converted[1], data_converted[2], data_converted[3]]
    return data_converted


def equationSolveForEndPoint(alpha, h, coord):
    if alpha < 90:
        newAlph = 90 - alpha

    if 100 > alpha >= 90:
        newAlph = 90 - alpha

    if alpha >= 180:
        newAlph = 90 - (alpha - 180)

    if 175 < alpha < 180:
        newAlph = 90 - (alpha - 180)

    h = float(h)
    x = coord[0]
    y = coord[1]
    x2 = x + (h * math.cos(math.radians(newAlph)))
    y2 = y + (h * math.sin(math.radians(newAlph)))

    return [x2, y2]


def getCoordsLst(valsLstMod, coord):
    coordLst = [[coord], [], [], []]
    yVals = [0.0]
    xVals = [0.0]
    for i in range(0, len(valsLstMod)):
        if i < 4:
            coord = equationSolveForEndPoint(valsLstMod[i][1], valsLstMod[i][0], coord)
            coordLst[0].append(coord)
            xVals.append(coord[0])
            yVals.append(coord[1])
        if 7 >= i >= 3:
            if i == 3:
                coordLst[1].append(coord[:])
            else:
                coord = equationSolveForEndPoint(valsLstMod[i][1], valsLstMod[i][0], coord)
                coordLst[1].append(coord)
            xVals.append(coord[0])
            yVals.append(coord[1])
        if 11 >= i >= 7:
            if i == 7:
                coordLst[2].append(coord[:])
            else:
                coord = equationSolveForEndPoint(valsLstMod[i][1], -valsLstMod[i][0], coord)
                coordLst[2].append(coord)
            xVals.append(coord[0])
            yVals.append(coord[1])

        if i >= 11:
            if i == 11:
                coordLst[3].append(coord[:])
            else:
                coord = equationSolveForEndPoint(valsLstMod[i][1], -valsLstMod[i][0], coord)
                coordLst[3].append(coord)
            xVals.append(coord[0])
            yVals.append(coord[1])
    coordLst1 = coordLst[:]

    return coordLst1


def translateDirectionToNumber(variable, val):
    if variable == 'rng':
        if val == 'W':
            return '2'
        elif val == 'E':
            return '1'
        else:
            return val
    elif variable == 'township':
        if val == 'S':
            return '2'
        elif val == 'N':
            return '1'
        else:
            return val
    elif variable == 'baseline':
        if val == 'U':
            return '2'
        elif val == 'S':
            return '1'
        else:
            return val
    elif variable == 'alignment':
        if val == 'SE':
            return '1'
        elif val == 'NE':
            return '2'
        elif val == 'SW':
            return '3'
        elif val == 'NW':
            return '4'
        else:
            return val


def translateNumberToDirection(variable, val):
    if variable == 'rng':
        if val == '2':
            return 'W'
        elif val == '1':
            return 'E'
    elif variable == 'township':
        if val == '2':
            return 'S'
        elif val == '1':
            return 'N'
    elif variable == 'baseline':
        if val == '2':
            return 'U'
        elif val == '1':
            return 'S'
    elif variable == 'alignment':
        if val == '1':
            return 'SE'
        elif val == '2':
            return 'NE'
        elif val == '3':
            return 'SW'
        elif val == '4':
            return 'NW'


def linesMain(coordLst, dirLst):
    newCoordLst = copy.deepcopy(coordLst)
    slopeLst = getSlope(coordLst)
    yIntLst = getYIntercept(coordLst, slopeLst)
    eqLst = [(slopeLst[i], yIntLst[i], dirLst[i]) for i in range(len(slopeLst))]
    coordLst1, cornersLst = connectAllCorners(slopeLst, yIntLst, coordLst)
    return eqLst, newCoordLst, cornersLst


def getYIntercept(coordLst, slopeLst):
    yIntLst = []
    count = 0
    for i in range(len(coordLst)):
        for j in range(len(coordLst[i]) - 1):
            x = coordLst[i][j][0]
            y = coordLst[i][j][1]
            m = slopeLst[count]

            yIntLst.append(y - (m * x))

            count += 1
    return yIntLst


def getSlope(coordLst):
    slopeLst1 = [[slopeFinder(i, j, coordLst) for j in range(len(coordLst[i]) - 1)] for i in range(len(coordLst))]
    slopeLst = [j for sub in slopeLst1 for j in sub]
    return slopeLst


def slopeFinder(i, j, coordLst):
    slopeValX = coordLst[i][j + 1][0] - coordLst[i][j][0]
    slopeValY = coordLst[i][j + 1][1] - coordLst[i][j][1]
    try:
        slopeVal = slopeValY / slopeValX
    except ZeroDivisionError:

        slopeVal = 0

    return slopeVal


def connectAllCorners(slopeLst, yIntLst, cLst):
    cnL = []
    # find the specific values for each corner and then create a sequential list of those coordinates
    corner11, corner12 = cLst[0][0], cLst[3][4]
    corner21, corner22 = cLst[0][4], cLst[1][0]
    corner31, corner32 = cLst[1][4], cLst[2][0]
    corner41, corner42 = cLst[2][4], cLst[3][0]
    cornersLst = [[corner11, corner12, 'SW'], [corner21, corner22, "SE"], [corner31, corner32, "NE"], [corner41, corner42, "NW"]]
    cnL.append([0, 0, 3, 4])
    cnL.append([0, 4, 1, 0])
    cnL.append([1, 4, 2, 0])
    cnL.append([2, 4, 3, 0])

    vals = [[0, 15], [3, 4], [7, 8], [11, 12]]
    sc = [[corner11, corner12], [corner21, corner22], [corner31, corner32], [corner41, corner42]]

    for i in range(4):
        # if the corner values are the same, ignore it
        if sc[i][0] == sc[i][1]:
            pass
        # if the corner values are not the same, find where the lines would intercept, and change the corners to that point
        else:
            # find the intercept between the two lines and change the value accordinly
            m1, m2 = slopeLst[vals[i][0]], slopeLst[vals[i][1]]
            y1, y2 = yIntLst[vals[i][0]], yIntLst[vals[i][1]]
            xy = getLineIntercept(m1, m2, y1, y2)
            for j in range(len(cLst)):
                for k in range(len(cLst[j])):
                    if cLst[j][k] == sc[i][0]:
                        cLst[j][k] = xy
                    if cLst[j][k] == sc[i][1]:
                        cLst[j][k] = xy
            cLst[cnL[i][0]][cnL[i][1]] = xy
            cLst[cnL[i][2]][cnL[i][3]] = xy

    corner11, corner12 = cLst[0][0], cLst[3][4]
    corner21, corner22 = cLst[0][4], cLst[1][0]
    corner31, corner32 = cLst[1][4], cLst[2][0]
    corner41, corner42 = cLst[2][4], cLst[3][0]
    cornersLst = [[corner11, corner12, 'SW'], [corner21, corner22, "SE"], [corner31, corner32, "NE"], [corner41, corner42, "NW"]]

    return cLst, cornersLst


# check for where the two lines intercept at
def getLineIntercept(m1, m2, b1, b2):
    if (m2 - m1) != 0:
        x = (b1 - b2) / (m2 - m1)
    else:
        x = (b1 - b2) / 1
    y = m1 * x + b1
    return [x, y]


def sqlConnect():
    conn = pyodbc.connect(
        "Driver={SQL Server};"
        "Server=DESKTOP-CMK3OJU\SQLEXPRESS;"
        "Database=UTRBDMSNET;"
        "Trusted_Connection = yes;")
    cursor = conn.cursor()
    return conn, cursor


def fixer(lst):
    fixed_lst = []
    for row in lst:
        fixed_lst.append(list(map(str, list(row))))
    return fixed_lst


class ZoomPan:
    def __init__(self):
        self.press = None
        self.cur_xlim = None
        self.cur_ylim = None
        self.x0 = None
        self.y0 = None
        self.x1 = None
        self.y1 = None
        self.xpress = None
        self.ypress = None

    def zoom_factory(self, ax, base_scale=2.):
        def zoom(event):

            cur_xlim = ax.get_xlim()
            cur_ylim = ax.get_ylim()

            xdata = event.xdata  # get event x location
            ydata = event.ydata  # get event y location

            if event.button == 'down':
                # deal with zoom in
                scale_factor = 1 / base_scale
            elif event.button == 'up':
                # deal with zoom out
                scale_factor = base_scale
            else:
                # deal with something that should never happen
                scale_factor = 1

            new_width = (cur_xlim[1] - cur_xlim[0]) * scale_factor
            new_height = (cur_ylim[1] - cur_ylim[0]) * scale_factor

            relx = (cur_xlim[1] - xdata) / (cur_xlim[1] - cur_xlim[0])
            rely = (cur_ylim[1] - ydata) / (cur_ylim[1] - cur_ylim[0])

            ax.set_xlim([xdata - new_width * (1 - relx), xdata + new_width * (relx)])
            ax.set_ylim([ydata - new_height * (1 - rely), ydata + new_height * (rely)])
            ax.figure.canvas.draw()

        fig = ax.get_figure()  # get the figure of interest
        fig.canvas.mpl_connect('scroll_event', zoom)

        return zoom

    def pan_factory(self, ax):
        def onPress(event):
            if event.inaxes != ax:
                return
            self.cur_xlim = ax.get_xlim()
            self.cur_ylim = ax.get_ylim()
            self.press = self.x0, self.y0, event.xdata, event.ydata
            self.x0, self.y0, self.xpress, self.ypress = self.press

        def onRelease(event):

            self.press = None
            ax.figure.canvas.draw()

        def onMotion(event):

            if self.press is None:
                return
            if event.inaxes != ax:
                return
            dx = event.xdata - self.xpress
            dy = event.ydata - self.ypress
            self.cur_xlim -= dx
            self.cur_ylim -= dy
            ax.set_xlim(self.cur_xlim)
            ax.set_ylim(self.cur_ylim)

            ax.figure.canvas.draw()

        fig = ax.get_figure()  # get the figure of interest

        # attach the call back
        fig.canvas.mpl_connect('button_press_event', onPress)
        fig.canvas.mpl_connect('button_release_event', onRelease)
        fig.canvas.mpl_connect('motion_notify_event', onMotion)

        # return the function
        return onMotion


# mainProcess()
