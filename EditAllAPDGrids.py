import re
import openpyxl
import os
import pandas as pd
import numpy as np
from glob import glob


def main():
    df = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length', 'Degrees', 'Minutes', 'Seconds', 'Alignment', 'North Reference', 'Concatenation'])
    currentDir = os.getcwd()
    # print(df)
    newPath = os.path.join(currentDir, 'Well Design')
    os.chdir(newPath)
    # filedirectory = os.listdir(newPath)
    # foo = [x[0] for x in os.walk(newPath)]
    output = glob(newPath + "/*/")
    counter = 0
    concLst = []
    plat_data = openpyxl.load_workbook("C:\Work\PlatGridNumbers.xlsm")
    pws = plat_data['PlatData']
    max_row_plat = pws.max_row


    # for k in range(3, max_row_plat + 1):---------------------------------
    #     section, township, township_dir, rng, rng_dir, baseline, side, length, degrees, minutes, seconds, alignment, nref, check, api, name = pws["A" + str(k)].value, pws["B" + str(k)].value, pws["C" + str(k)].value, pws["D" + str(k)].value, pws["E" + str(k)].value, pws[
    #         "F" + str(k)].value, pws["G" + str(k)].value, pws["H" + str(k)].value, pws["I" + str(k)].value, pws["J" + str(k)].value, pws["K" + str(k)].value, pws["L" + str(k)].value, pws["M" + str(k)].value, pws["N" + str(k)].value, pws["O" + str(k)].value, pws["P" + str(k)].value
    #     conc = str(section) + str(township) + str(township_dir) + str(rng) + str(rng_dir) + str(baseline) + str(side)
    #     if conc not in concLst:
    #         concLst.append(conc)
    #         # print(conc)
    #         new_row = {'Section': str(section), 'Township': str(township), 'Township Direction': str(township_dir),
    #                    'Range': str(rng), 'Range Direction': str(rng_dir), 'Baseline': str(baseline), 'Side': str(side),
    #                    'Length': str(length), 'Degrees': str(degrees), 'Minutes': str(minutes),
    #                    'Seconds': str(seconds), 'Alignment': str(alignment),
    #                    'North Reference': str(nref), 'Concatenation': str(conc)}
    #         # print(new_row)
    #         df = df.append(new_row, ignore_index=True)
    print(df)
    findExistingGoodclosure(df)
    known_conc = []
    all_sections = []
    for i in output:
        for j in os.listdir(i):
            if 'xl' in j and 'casing' in j.lower() and 'review' in j.lower():
                file_path = os.path.join(newPath, i, j)
                try:
                    apdWorkbook = openpyxl.load_workbook(file_path, keep_vba=False, data_only=True)
                    try:
                        apdWS = apdWorkbook['Grid Numbers']
                        apdCR = apdWorkbook['Casing Review']
                        apdWSMaxRow = apdWS.max_row
                        # apdSHL = apdWorkbook['SHL Section']
                        # apdBHL1 = apdWorkbook['BHL Section 1']
                        # apdBHL2 = apdWorkbook['BHL Section 2']
                        # apdBHL3 = apdWorkbook['BHL Section 3']
                        # shl_cells = ['CC22', 'CD22']
                        # bhl_1cells = ['CC22', 'CD22']

                        # if abs(apdSHL[shl_cells[0]].value) < 2 and abs(apdSHL[shl_cells[1]].value) < 2:
                        #     data_set = [apdSHL['N7'].value, apdSHL['O7'].value, apdSHL['P7'].value, apdSHL['Q7'].value, apdSHL['R7'].value, apdSHL['S7'].value, apdCR['B6'].value]
                        #
                        #     # data_set[2] = translateDirectionToNumber('township', data_set[2])
                        #     #
                        #     # data_set[4] = translateDirectionToNumber('rng', data_set[4])
                        #     #
                        #     # data_set[5] = translateDirectionToNumber('baseline', data_set[5])
                        #     # data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                        #
                        #     conc = "".join([str(k) for k in data_set[:-1]])
                        #
                        #     if conc not in known_conc:
                        #         print(0, 'conc', conc, data_set)
                        #         known_conc.append(conc)
                        #         section_data = gatherDataCells(apdSHL, 'shl', data_set)
                        #         # all_sections = all_sections + section_data
                        #         df = addToDF(section_data, df)
                        #         print(df)
                        #
                        # if abs(apdBHL1[shl_cells[0]].value) < 2 and abs(apdBHL1[shl_cells[1]].value) < 2:
                        #     data_set = [apdBHL1['N8'].value, apdBHL1['O8'].value, apdBHL1['P8'].value, apdBHL1['Q8'].value, apdBHL1['R8'].value, apdBHL1['S8'].value, apdCR['B6'].value]
                        #     data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                        #     conc = "".join([str(k) for k in data_set[:-1]])
                        #     if conc not in known_conc:
                        #         print(1, 'conc', conc, data_set)
                        #         known_conc.append(conc)
                        #         section_data = gatherDataCells(apdBHL1, 'bhl', data_set)
                        #         df = addToDF(section_data, df)
                        #         # all_sections = all_sections + section_data
                        #
                        # if abs(apdBHL2[shl_cells[0]].value) < 2 and abs(apdBHL2[shl_cells[1]].value) < 2:
                        #     data_set = [apdBHL2['N9'].value, apdBHL2['O9'].value, apdBHL2['P9'].value, apdBHL2['Q9'].value, apdBHL2['R9'].value, apdBHL2['S9'].value, apdCR['B6'].value]
                        #     data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                        #     conc = "".join([str(k) for k in data_set[:-1]])
                        #     if conc not in known_conc:
                        #         print(2, 'conc', conc, data_set)
                        #         known_conc.append(conc)
                        #         section_data = gatherDataCells(apdBHL2, 'bhl', data_set)
                        #         df = addToDF(section_data, df)
                        #         # all_sections = all_sections + section_data
                        #     # known_conc.append([apdBHL2['N9'].value,  apdBHL2['O9'].value, apdBHL2['P9'].value, apdBHL2['Q9'].value, apdBHL2['R9'].value, "\t|\t", apdCR['B6'].value])
                        #
                        # if abs(apdBHL3[shl_cells[0]].value) < 2 and abs(apdBHL3[shl_cells[1]].value) < 2:
                        #     data_set = [apdBHL3['N10'].value, apdBHL3['O10'].value, apdBHL3['P10'].value, apdBHL3['Q10'].value, apdBHL3['R10'].value, apdBHL2['S10'].value, apdCR['B6'].value]
                        #     data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                        #     conc = "".join([str(k) for k in data_set[:-1]])
                        #     if conc not in known_conc:
                        #         print(3, 'conc', conc, data_set)
                        #         known_conc.append(conc)
                        #         section_data = gatherDataCells(apdBHL3, 'bhl', data_set)
                        #         df = addToDF(section_data, df)
                        #         # all_sections = all_sections + section_data
                        #     # known_conc.append([apdBHL3['N10'].value,  apdBHL3['O10'].value, apdBHL3['P10'].value, apdBHL3['Q10'].value, apdBHL3['R10'].value, "\t|\t", apdCR['B6'].value])

                        # api, name = apdCR['B6'], apdCR['B5']
                        # for k in range(3, apdWSMaxRow+1):
                        #     section, township, township_dir, rng, rng_dir, baseline, side, length, degrees, minutes, seconds, alignment, nref, check, api, name = apdWS["A" + str(k)].value, apdWS["B" + str(k)].value, apdWS["C" + str(k)].value, apdWS["D" + str(k)].value, apdWS["E" + str(k)].value,apdWS["F" + str(k)].value, apdWS["G" + str(k)].value, apdWS["H" + str(k)].value, apdWS["I" + str(k)].value, apdWS["J" + str(k)].value,apdWS["K" + str(k)].value, apdWS["L" + str(k)].value, apdWS["M" + str(k)].value, apdWS["N" + str(k)].value, apdWS["O" + str(k)].value,apdWS["P" + str(k)].value
                        #     conc = str(section) + str(township)+ str(township_dir)+ str(rng)+ str(rng_dir)+ str(baseline)+ str(side)
                        #     if conc == 4922111 or conc == '4922111':
                        #         print('api here', j)
                        #     if conc not in concLst:
                        #         concLst.append(conc)
                        #
                        #         new_row = {'Section': str(section), 'Township': str(township), 'Township Direction': str(township_dir), 'Range': str(rng), 'Range Direction': str(rng_dir), 'Baseline': str(baseline), 'Side': str(side), 'Length': str(length), 'Degrees': str(degrees),
                        #                    'Minutes': str(minutes), 'Seconds': str(seconds), 'Alignment': str(alignment),
                        #                    'North Reference': str(nref), 'Checked': str(check), 'API': str(api), 'WellName': str(name), 'Concatenation': str(conc)}
                        #         df = df.append(new_row, ignore_index=True)
                        # print(section, township, township_dir, rng, rng_dir, baseline, side, length, degrees, minutes, seconds, alignment, nref, check, api, name, conc, j)
                        # section, township, township_dir, rng, rng_dir, baseline, side, length, degrees, minutes, seconds, alignment, nref, check, api, name =

                        # print(j, counter)
                        # counter +=1
                    except KeyError:
                        pass
                except:
                    pass

    print(df)
    df.to_csv('parsed_closure.csv')


def findExistingGoodclosure(df):
    known_conc = []
    all_sections = []
    for i in output:
        for j in os.listdir(i):
            if 'xl' in j and 'casing' in j.lower() and 'review' in j.lower():
                file_path = os.path.join(newPath, i, j)
                try:
                    apdWorkbook = openpyxl.load_workbook(file_path, keep_vba=False, data_only=True)
                    try:
                        apdWS = apdWorkbook['Grid Numbers']
                        apdCR = apdWorkbook['Casing Review']
                        apdSHL = apdWorkbook['SHL Section']
                        apdBHL1 = apdWorkbook['BHL Section 1']
                        apdBHL2 = apdWorkbook['BHL Section 2']
                        apdBHL3 = apdWorkbook['BHL Section 3']
                        shl_cells = ['CC22', 'CD22']
                        bhl_1cells = ['CC22', 'CD22']
                        apdWSMaxRow = apdWS.max_row
                        if abs(apdSHL[shl_cells[0]].value) < 2 and abs(apdSHL[shl_cells[1]].value) < 2:
                            data_set = [apdSHL['N7'].value, apdSHL['O7'].value, apdSHL['P7'].value, apdSHL['Q7'].value, apdSHL['R7'].value, apdSHL['S7'].value, apdCR['B6'].value]

                            # data_set[2] = translateDirectionToNumber('township', data_set[2])
                            #
                            # data_set[4] = translateDirectionToNumber('rng', data_set[4])
                            #
                            # data_set[5] = translateDirectionToNumber('baseline', data_set[5])
                            # data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])

                            conc = "".join([str(k) for k in data_set[:-1]])

                            if conc not in known_conc:
                                print(0, 'conc', conc, data_set)
                                known_conc.append(conc)
                                section_data = gatherDataCells(apdSHL, 'shl', data_set)
                                # all_sections = all_sections + section_data
                                df = addToDF(section_data, df)
                                print(df)

                        if abs(apdBHL1[shl_cells[0]].value) < 2 and abs(apdBHL1[shl_cells[1]].value) < 2:
                            data_set = [apdBHL1['N8'].value, apdBHL1['O8'].value, apdBHL1['P8'].value, apdBHL1['Q8'].value, apdBHL1['R8'].value, apdBHL1['S8'].value, apdCR['B6'].value]
                            data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                            conc = "".join([str(k) for k in data_set[:-1]])
                            if conc not in known_conc:
                                print(1, 'conc', conc, data_set)
                                known_conc.append(conc)
                                section_data = gatherDataCells(apdBHL1, 'bhl', data_set)
                                df = addToDF(section_data, df)
                                # all_sections = all_sections + section_data

                        if abs(apdBHL2[shl_cells[0]].value) < 2 and abs(apdBHL2[shl_cells[1]].value) < 2:
                            data_set = [apdBHL2['N9'].value, apdBHL2['O9'].value, apdBHL2['P9'].value, apdBHL2['Q9'].value, apdBHL2['R9'].value, apdBHL2['S9'].value, apdCR['B6'].value]
                            data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                            conc = "".join([str(k) for k in data_set[:-1]])
                            if conc not in known_conc:
                                print(2, 'conc', conc, data_set)
                                known_conc.append(conc)
                                section_data = gatherDataCells(apdBHL2, 'bhl', data_set)
                                df = addToDF(section_data, df)
                                # all_sections = all_sections + section_data
                            # known_conc.append([apdBHL2['N9'].value,  apdBHL2['O9'].value, apdBHL2['P9'].value, apdBHL2['Q9'].value, apdBHL2['R9'].value, "\t|\t", apdCR['B6'].value])

                        if abs(apdBHL3[shl_cells[0]].value) < 2 and abs(apdBHL3[shl_cells[1]].value) < 2:
                            data_set = [apdBHL3['N10'].value, apdBHL3['O10'].value, apdBHL3['P10'].value, apdBHL3['Q10'].value, apdBHL3['R10'].value, apdBHL2['S10'].value, apdCR['B6'].value]
                            data_set[2], data_set[4], data_set[5] = translateDirectionToNumber('township', data_set[2]), translateDirectionToNumber('rng', data_set[4]), translateDirectionToNumber('baseline', data_set[5])
                            conc = "".join([str(k) for k in data_set[:-1]])
                            if conc not in known_conc:
                                print(3, 'conc', conc, data_set)
                                known_conc.append(conc)
                                section_data = gatherDataCells(apdBHL3, 'bhl', data_set)
                                df = addToDF(section_data, df)
                                # all_sections = all_sections + section_data
                            # known_conc.append([apdBHL3['N10'].value,  apdBHL3['O10'].value, apdBHL3['P10'].value, apdBHL3['Q10'].value, apdBHL3['R10'].value, "\t|\t", apdCR['B6'].value])



def addToDF(lst, df):
    # print(lst)
    for i in lst:
        # print(i)
        new_row = {'Section': str(i[0]), 'Township': str(i[1]), 'Township Direction': str(i[2]), 'Range': str(i[3]),
                   'Range Direction': str(i[4]), 'Baseline': str(i[5]), 'Side': str(i[6]), 'Length': str(i[7]), 'Degrees': str(i[8]),
                   'Minutes': str(i[9]), 'Seconds': str(i[10]), 'Alignment': str(i[11]),
                   'North Reference': str(i[12]), 'Concatenation': str(i[13])}
        # print(new_row)
        df = df.append(new_row, ignore_index=True)
    return df
    # df.to_csv('grid_data_all.csv')


    # for i in range(len(filedirectory)):
    #     findFilesRecursion(filedirectory[i])
    # for i in os.listdir(newPath):
    #     print(i)
    #     if os.path.isfile(os.path.join(newPath, i)) and 'Casing Review' in i and 'xlsx' in i:
    #         print(i)
    #         apdWorkbook = openpyxl.load_workbook(i)
    #         apdWS = apdWorkbook['Grid Numbers']
    # known_conc = [list(t) for t in set(tuple(element) for element in known_conc)]



def gatherDataCells(ws, label, tsr_data):

    ref = 'T'
    conc_lst_all = []
    lstValues = []
    lstCoords = []
    firstRowValue = -1
    azimuthVals = []
    lstValuesFlatten = []
    # sheet = gridbook.active
    max_row = ws.max_row
    dirLst = ['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2',
              'East-Up2', 'East-Up1', 'East-Down1', 'East-Down2',
              'North-Left2', 'North-Left1', 'North-Right1', 'North-Right2',
              'South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']

    writeCoordLstSHL = [['West-Up2', 'B25', 'B26', 'B27', 'B28', 'B29'],
                        ['West-Up1', 'B33', 'B34', 'B35', 'B36', 'B37'],
                        ['West-Down1', 'B41', 'B42', 'B43', 'B44', 'B45'],
                        ['West-Down2', 'B49', 'B50', 'B51', 'B52', 'B53'],
                        ['East-Up2', 'W25', 'W26', 'W27', 'W28', 'W29'],
                        ['East-Up1', 'W33', 'W34', 'W35', 'W36', 'W37'],
                        ['East-Down1', 'W41', 'W42', 'W43', 'W44', 'W45'],
                        ['East-Down2', 'W49', 'W50', 'W51', 'W52', 'W53'],
                        ['North-Left2', 'E18', 'E19', 'E20', 'E21', 'E22'],
                        ['North-Left1', 'K18', 'K19', 'K20', 'K21', 'K22'],
                        ['North-Right1', 'O18', 'O19', 'O20', 'O21', 'O22'],
                        ['North-Right2', 'U18', 'U19', 'U20', 'U21', 'U22'],
                        ['South-Left2', 'E57', 'E58', 'E59', 'E60', 'E61'],
                        ['South-Left1', 'K57', 'K58', 'K59', 'K60', 'K61'],
                        ['South-Right1', 'O57', 'O58', 'O59', 'O60', 'O61'],
                        ['South-Right2', 'U57', 'U58', 'U59', 'U60', 'U61']]

    writeCoordLstBHL = [['West-Up2', 'B25', 'B26', 'B27', 'B28', 'B29'],
                        ['West-Up1', 'B33', 'B34', 'B35', 'B36', 'B37'],
                        ['West-Down1', 'B42', 'B43', 'B44', 'B45', 'B46'],
                        ['West-Down2', 'B50', 'B51', 'B52', 'B53', 'B54'],
                        ['East-Up2', 'W25', 'W26', 'W27', 'W28', 'W29'],
                        ['East-Up1', 'W33', 'W34', 'W35', 'W36', 'W37'],
                        ['East-Down1', 'W42', 'W43', 'W44', 'W45', 'W46'],
                        ['East-Down2', 'W50', 'W51', 'W52', 'W53', 'W54'],
                        ['North-Left2', 'E18', 'E19', 'E20', 'E21', 'E22'],
                        ['North-Left1', 'K18', 'K19', 'K20', 'K21', 'K22'],
                        ['North-Right1', 'O18', 'O19', 'O20', 'O21', 'O22'],
                        ['North-Right2', 'U18', 'U19', 'U20', 'U21', 'U22'],
                        ['South-Left2', 'E57', 'E58', 'E59', 'E60', 'E61'],
                        ['South-Left1', 'K57', 'K58', 'K59', 'K60', 'K61'],
                        ['South-Right1', 'O57', 'O58', 'O59', 'O60', 'O61'],
                        ['South-Right2', 'U57', 'U58', 'U59', 'U60', 'U61']]

    counter = 0
    section_dat = []

    if label == 'shl':
        for i in range(len(writeCoordLstSHL)):
            dataset_measurements = []
            for j in range(1, len(writeCoordLstSHL[i])):
                cell_val = ws[writeCoordLstSHL[i][j]].value
                dataset_measurements.append(cell_val)
            conc_lst = "".join([str(k) for k in tsr_data[:-1]]) + dirLst[i]
            ref = ws['Q13'].value
            if ref == 1:
                ref = 'T'
            elif ref == 2:
                ref == 'G'
            output_to_print = tsr_data[:-1] + [dirLst[i]] + dataset_measurements + [ref] + [conc_lst]
            section_dat.append(output_to_print)
            counter += 1

    elif label == 'bhl':
        for i in range(len(writeCoordLstBHL)):
            dataset_measurements = []
            for j in range(1, len(writeCoordLstBHL[i])):
                cell_val = ws[writeCoordLstBHL[i][j]].value
                dataset_measurements.append(cell_val)
            conc_lst = "".join([str(k) for k in tsr_data[:-1]]) + dirLst[i]
            ref = ws['Q13'].value
            if ref == 1:
                ref = 'T'
            elif ref == 2:
                ref == 'G'
            output_to_print = tsr_data[:-1] + [dirLst[i]] + dataset_measurements + [ref] + [conc_lst]
            section_dat.append(output_to_print)
            counter += 1
    return section_dat


def translateNumberToDirection(variable, val):
    if variable == 'rng':
        if val == '2':
            return 'W'
        elif val == '1':
            return 'E'
    elif variable == 'township':
        if val == '2':
            return 'S'
        elif val == '1':
            return 'N'
    elif variable == 'baseline':
        if val == '2':
            return 'U'
        elif val == '1':
            return 'S'
    elif variable == 'alignment':
        if val == '1':
            return 'SE'
        elif val == '2':
            return 'NE'
        elif val == '3':
            return 'SW'
        elif val == '4':
            return 'NW'


def translateDirectionToNumber(variable, val):
    # print(variable, val)
    if variable == 'rng':
        if val == 'W':
            return '2'
        elif val == 'E':
            return '1'
        else:
            return val
        if isinstance(val, int) or isinstance(val, float):
            return val
    elif variable == 'township':
        if isinstance(val, int) or isinstance(val, float):
            return val
        if val == 'S':
            return '2'
        elif val == 'N':
            return '1'
        else:
            return val
    elif variable == 'baseline':
        if isinstance(val, int) or isinstance(val, float):
            return val
        if val == 'U' or val.lower() == 'uintah' or val.lower() == 'uinta':
            return '2'
        elif val == 'S' or val.lower() == 'salt lake' or val.lower() == 'saltlake':
            return '1'
        else:
            return val

    elif variable == 'alignment':
        if isinstance(val, int) or isinstance(val, float):
            return val
        if val == 'SE':
            return '1'
        elif val == 'NE':
            return '2'
        elif val == 'SW':
            return '3'
        elif val == 'NW':
            return '4'
        else:
            return val

    return val

def findFilesRecursion(val):
    if val.isdir():
        print(val)

        # townshipRangePath = os.path.join(currentDir, 'GridNumbers.xlsx')
        # gridWorkbook = openpyxl.load_workbook(townshipRangePath)
        # tsrWS = gridWorkbook['Sheet2']
        #
        # apdWSMinRow = 3
        # apdWSMaxRow = apdWS.max_row
        # apdWSMaxCol = apdWS.max_column
        #
        # gridWSMinRow = 2
        # grdWSMaxRow = tsrWS.max_row
        # grdWSMaxCol = tsrWS.max_column


main()
