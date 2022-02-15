import openpyxl
import os
import shutil
from glob import glob
import pandas as pd
import numpy as np
import ModuleAgnostic as ma


def main():
    # parseForUpdatedApds()
    parseThroughAPDS()


def parseThroughAPDS():
    path = "C:\\Work\\newapds\\PlatGridNumbers.csv"
    df_all = pd.read_csv(path, encoding="ISO-8859-1")
    lst = df_all.to_numpy().tolist()
    reeditConcForPlatGridNumbers(lst)
    lst_conc_data = [i[13] for i in lst]
    dir_lst = [['West-Up2', 'West-Up1', 'West-Down1', 'West-Down2'],
               ['East-Up2', 'East-Up1', 'East-Down1', 'East-Down2'],
               ['North-Left2', 'North-Left1', 'North-Right1', 'North-Right2'],
               ['South-Left2', 'South-Left1', 'South-Right1', 'South-Right2']]
    files = glob("C:\\Work\\newapds" + '/**/', recursive=True)

    for i in files:
        # print()
        for j in os.listdir(i):
            if 'casing' in j.lower():
                file_path = os.path.join(i, j)
                wb = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)
                grid_ws = wb['Grid Numbers']
                wsSHL, wsBHL1, wsBHL2, wsBHL3 = wb['SHL Section'], wb['BHL Section 1'], wb['BHL Section 2'], wb['BHL Section 3']
                # checkExcelData(grid_ws, lst_conc_data)
                # checkEachPage(wsSHL, lst_conc_data)


def reeditConcForPlatGridNumbers(lst):
    df_base = pd.DataFrame(columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length',
                                    'Degrees', 'Minutes', 'Seconds', 'Alignment', 'Concatenation', 'Version'])
    for i in range(len(lst)):
        lst[i][-2] = str(ma.reTranslateData(lst[i])) + lst[i][6]

    # df_test = [{'Section': i[0], 'Township': int(float(i[1])), 'Township Direction': i[2], 'Range': int(float(i[3])),
    #             'Range Direction': i[4], 'Baseline': i[5], 'Side': float(i[6]), 'Length': float(i[7]), 'Degrees': float(i[8]), "Minutes": float(i[9]),
    #             'Seconds': i[10], 'Alignment': i[11],'Concatenation': i[12],'Version': i[13]} for i in lst]
    # df = pd.DataFrame(df_test, columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Concatenation', 'Version'])

    df_test = [{'Section': i[0], 'Township': int(float(i[1])), 'Township Direction': i[2], 'Range': int(float(i[3])),
                'Range Direction': i[4], 'Baseline': i[5], 'Side': i[6], 'Length': i[7], 'Degrees': i[8], 'Minutes': i[9], 'Seconds': i[10], 'Direction': i[11], 'North Reference':i[12], 'Conc': i[13], 'Version':i[14]} for i in lst]
    df = pd.DataFrame(df_test, columns=['Section', 'Township', 'Township Direction', 'Range', 'Range Direction', 'Baseline', 'Side', 'Length', 'Degrees', 'Minutes', 'Seconds', 'Direction', 'North Reference', 'Conc', 'Version'])
    # print(df)
    # # for i in lst:
    # #     new_row = {'Section': i[0],
    # #                'Township': int(float(i[1])),
    # #                'Township Direction': i[2],
    # #                'Range': int(float(i[3])),
    # #                'Range Direction': i[4],
    # #                'Baseline': i[5],
    # #                'Side': i[6],
    # #                'Length': i[7],
    # #                'Degrees': i[8],
    # #                'Minutes': i[9],
    # #                'Seconds': i[10],
    # #                'Alignment': i[11],
    # #                'Concatenation': i[12],
    # #                'Version': i[13]}
    # #     df_base = df_base.append(new_row, ignore_index=True)
    # # print(df_base)
    return lst


def checkExcelData(ws, lst_conc):
    max_row = ws.max_row
    alpha = list('ABCDEFGHIJKLMNOPQ')
    for i in range(3, max_row + 1):
        conc_val = [ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value, ws['E' + str(i)].value, ws['F' + str(i)].value, ws['G' + str(i)].value]
        tsr_data = [ws[r + str(i)].value for r in alpha]
        # tsr_data = ['N7', 'O7', 'P7', 'Q7', 'R7', 'S7']
        if conc_val not in lst_conc:
            print(tsr_data)
            print('new1', conc_val)

def checkEachPage(ws, lst_conc_data):
    tsr_data = ['N7', 'O7', 'P7', 'Q7', 'R7', 'S7']
    tsr_data = [ws[i].value for i in tsr_data]
    tsr_data_conc = "".join([str(i) for i in tsr_data])
    if tsr_data_conc+'West-Up2' not in lst_conc_data:
        print('new', tsr_data_conc)
    # print(tsr_data)
    # writeCoordLstSHL = [['West-Up2', 'D25', 'D26', 'D27', 'D28', 'D29'],
    #                     ['West-Up1', 'D33', 'D34', 'D35', 'D36', 'B37'],
    #                     ['West-Down1', 'D41', 'D42', 'D43', 'D44', 'D45'],
    #                     ['West-Down2', 'D49', 'D50', 'D51', 'D52', 'D53'],
    #                     ['East-Up2', 'U25', 'U26', 'U27', 'U28', 'U29'],
    #                     ['East-Up1', 'U33', 'U34', 'U35', 'U36', 'U37'],
    #                     ['East-Down1', 'U41', 'U42', 'U43', 'U44', 'U45'],
    #                     ['East-Down2', 'U49', 'U50', 'U51', 'U52', 'U53'],
    #                     ['North-Left2', 'F18', 'F19', 'F20', 'F21', 'F22'],
    #                     ['North-Left1', 'L18', 'L19', 'L20', 'L21', 'L22'],
    #                     ['North-Right1', 'P18', 'P19', 'P20', 'P21', 'P22'],
    #                     ['North-Right2', 'V18', 'V19', 'V20', 'V21', 'V22'],
    #                     ['South-Left2', 'F57', 'F58', 'F59', 'F60', 'F61'],
    #                     ['South-Left1', 'L57', 'L58', 'L59', 'L60', 'L61'],
    #                     ['South-Right1', 'P57', 'P58', 'P59', 'P60', 'P61'],
    #                     ['South-Right2', 'V57', 'V58', 'V59', 'V60', 'V61']]
    writeCoordLstBHL = [['West-Up2', 'B25', 'B26', 'B27', 'B28', 'B29'],
                        ['West-Up1', 'B33', 'B34', 'B35', 'B36', 'B37'],
                        ['West-Down1', 'B42', 'B43', 'B44', 'B45', 'B46'],
                        ['West-Down2', 'B50', 'B51', 'B52', 'B53', 'B54'],
                        ['East-Up2', 'W25', 'W26', 'W27', 'W28', 'W29'],
                        ['East-Up1', 'W33', 'W34', 'W35', 'W36', 'W37'],
                        ['East-Down1', 'W42', 'W43', 'W44', 'W45', 'W46'],
                        ['East-Down2', 'W50', 'W51', 'W52', 'W53', 'W54'],
                        ['North-Left2', 'E18', 'E19', 'E20', 'E21', 'E22'],
                        ['North-Left1', 'K18', 'K19', 'K20', 'K21', 'K22'],
                        ['North-Right1', 'O18', 'O19', 'O20', 'O21', 'O22'],
                        ['North-Right2', 'U18', 'U19', 'U20', 'U21', 'U22'],
                        ['South-Left2', 'E57', 'E58', 'E59', 'E60', 'E61'],
                        ['South-Left1', 'K57', 'K58', 'K59', 'K60', 'K61'],
                        ['South-Right1', 'O57', 'O58', 'O59', 'O60', 'O61'],
                        ['South-Right2', 'U57', 'U58', 'U59', 'U60', 'U61']]
    for i in range(len(writeCoordLstBHL)):
        line_lst = [ws[writeCoordLstBHL[i][j]].value for j in range(1, len(writeCoordLstBHL[i]))]
        # print(tsr_data, writeCoordLstBHL[i][0], line_lst)
        # for j in range(1, len(writeCoordLstBHL[i])):
        #     print(ws[writeCoordLstBHL[i][j]].value)



def parseForUpdatedApds():
    path = "J:\\Well design"
    file_paths_new = parseAllFoldersForString(path)
    for item in file_paths_new:
        if 'xlsx' in item.lower():
            copyToNewLocation(item)
    pass


def parseAllFoldersForString(new_path):
    files = glob(new_path + '/**/', recursive=True)
    created_data = []
    original_time = 1636398382.4097853
    file_paths_new = []
    for i in files:
        for j in os.listdir(i):
            file_path = os.path.join(i, j)
            created = os.path.getctime(file_path)
            if created > original_time:
                file_paths_new.append(file_path)
    return file_paths_new


def copyToNewLocation(path):
    shutil.copy(path, "C:\\Work\\newapds")


main()
